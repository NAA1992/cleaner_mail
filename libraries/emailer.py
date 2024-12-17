import sys
from pathlib import Path
sys.path.append(str(Path(__file__).resolve().parent.parent))

from libraries.common_funcs import CustomFormatter, set_custom_logger

from functools import partial
import os
import re
import openpyxl
import imaplib
import email
import concurrent.futures
import logging
import codecs
from collections import defaultdict
from email.header import decode_header
from email.utils import parseaddr
from pathlib import Path
from imapclient import imap_utf7

logger = set_custom_logger(name_logger="DEBUGGER", level_logger="DEBUG")


from typing import Tuple, Union, Optional, Dict, List, Any

class EmailDataProcessor:
    #MARK: EmailDataProcessor
    def __init__(self,
                imap_server:str,
                email_user:str,
                email_password:str,
                output_excel_file:str,
                input_excel_file:str="",
                exclude_folders:List[str]=[]):
        self.imap_server = imap_server
        self.email_user = email_user
        self.email_password = email_password
        self.output_excel_file = Path(output_excel_file).resolve()
        self.input_excel_file = Path(input_excel_file).resolve() if input_excel_file else None
        self.exclude_folders = set(exclude_folders or [])



    def init_self_logger(self, logger_name:str):
        #MARK: init_self_logger
        self.logger = set_custom_logger(name_logger=logger_name)



    def parse_email_address(self,
                            raw_from
                            ) -> Tuple[str, str]:
        #MARK: parse_email_address
        name, addr = parseaddr(raw_from)
        decoded_name = decode_header(name)
        name = ''.join(
            part.decode(charset or 'utf-8') if isinstance(part, bytes) else part
            for part, charset in decoded_name
        )
        return (name, addr)



    def decode_subject(self,
                       subject:str,
                       ) -> str:
        #MARK: decode_subject
        # Пробуем декодировать как заголовок письма
        decoded_parts = decode_header(subject or "")
        decoded_subject = ''.join(
            part.decode(charset or 'utf-8', errors='replace') if isinstance(part, bytes) else part
            for part, charset in decoded_parts
        )

        return decoded_subject


    def extract_attachments(self,
                            msg):
        #MARK: extract_attachments
        attachments = []
        for part in msg.walk():
            if part.get_content_disposition() == 'attachment':
                filename = part.get_filename()
                if filename:
                    decoded_filename = self.decode_subject(filename)
                    attachments.append(decoded_filename)
        return ", ".join(attachments)



    def fetch_emails_from_folder(self,
                                folder_name:bytes):
        #MARK: fetch_emails_from_folder
        emails = []
        try:
            folder_decoded = folder_name.decode().split('"')[-2]
            folder_decoded_rus = imap_utf7.decode(folder_name).split('"')[-2]
            #if any(exclude in folder_decoded_rus for exclude in self.exclude_folders):
            #    print(f"Папка находится среди исключений: {folder_decoded_rus}")
            #    return
            print(f"В работе папка {folder_decoded_rus}")
            with imaplib.IMAP4_SSL(self.imap_server) as mail:
                mail.login(self.email_user, self.email_password)
                mail.select(f'"{folder_decoded}"', readonly=True)
                status, messages = mail.search(None, 'ALL')
                if status != 'OK':
                    raise Exception(f"Не удалось извлечь письма из папки {folder_decoded_rus}")
                email_ids = messages[0].split()
                email_ids_size = len(email_ids)
                print(f"Количество писем в папке {folder_decoded_rus}: {email_ids_size}")
                for email_id in email_ids:
                    status, msg_data = mail.fetch(email_id, '(RFC822)')
                    for response_part in msg_data:
                        if isinstance(response_part, tuple):
                            msg = email.message_from_bytes(response_part[1])
                            name_sender, email_address = self.parse_email_address(msg['From'])
                            subject = self.decode_subject(msg['Subject'])
                            size = len(response_part[1])
                            attachments = self.extract_attachments(msg)
                            date = msg['Date']
                            emails.append((email_address, name_sender, subject, date, size, attachments))
                            if len(emails) >= 100:
                                print(f"Папка {folder_decoded_rus}: достигнута обработка 100 писем, записываем их в файл")
                                self.append_to_excel(emails)
                                emails.clear()
            if emails:
                print(f"Папка {folder_decoded_rus}: достигнут конец обработки папки, записываем письма в файл")
                self.append_to_excel(emails)
        except Exception as e:
            try:
                print(f"Произошла ошибка в папке {folder_decoded_rus}")
            except Exception:
                pass
            print(f"Произошла ошибка: {str(e)}")
            pass



    def fetch_emails(self):
        #MARK: fetch_emails
        with imaplib.IMAP4_SSL(self.imap_server) as mail:
            mail.login(self.email_user, self.email_password)
            status, folders = mail.list()
            if status != 'OK':
                raise Exception("Failed to retrieve folders")
            print("Получили список папок на обработку")
            # folders = [folder_name.decode().split('"')[-2] for folder_name in folders]
            with concurrent.futures.ProcessPoolExecutor() as executor:
                all_emails = executor.map(self.fetch_emails_from_folder, folders)





    def append_to_excel(self,
                        emails):
        #MARK: append_to_excel
        lock_file = f"{self.output_excel_file}.lock"
        with open(lock_file, 'w') as lock:
            try:
                wb = openpyxl.load_workbook(self.output_excel_file)
            except FileNotFoundError:
                wb = openpyxl.Workbook()

            headers = ["Email", "Name_Address", "Subject", "Date", "Size (bytes)", "Attachments"]
            ws = wb.active
            if ws.title != "Raw Data":
                ws = wb.create_sheet("Raw Data")

            for content in emails:
                ws.append(list(content))
            wb.save(self.output_excel_file)
        os.remove(lock_file)



    def group_by_email(self,
                       emails
                       ) -> Dict[str, int]:
        #MARK: group_by_email
        email_summary = defaultdict(lambda: {'count': 0, 'size': 0})
        for email, _, size in emails:
            email_summary[email]['count'] += 1
            email_summary[email]['size'] += size
        return email_summary


    def group_by_domain(self,
                        emails
                        ) -> Dict[str, Any]:
        #MARK: group_by_domain
        domain_summary = defaultdict(lambda: {'count': 0, 'unique_emails': set(), 'size': 0})
        for email, _, size in emails:
            domain = '.'.join(email.split('@')[-1].split('.')[-2:])
            domain_summary[domain]['count'] += 1
            domain_summary[domain]['unique_emails'].add(email)
            domain_summary[domain]['size'] += size
        return domain_summary



    def create_excel(self,
                    emails
                    ):
        #MARK: create_excel
        wb = openpyxl.Workbook()

        ws1 = wb.active
        ws1.title = "By Email"
        #ws1.append(["Email", "Count", "Total Size (bytes)"])
        #for email, data in self.group_by_email(emails).items():
        #    ws1.append([email, data['count'], data['size']])

        ws2 = wb.create_sheet("By Domain")
        ws2.append(["Domain", "Count", "Unique Emails", "Total Size (bytes)"])
        #for domain, data in self.group_by_domain(emails).items():
        #    ws2.append([domain, data['count'], len(data['unique_emails']), data['size']])

        ws3 = wb.create_sheet("Raw Data")
        ws3.append(["folder_name", "email_address", "name_address", "subject", "size"])
        for folder_name, email_address, name_address, subject, size in emails:
            ws3.append([folder_name, email_address, name_address, subject, size])

        wb.save(self.output_excel_file)



    def load_emails_from_excel(self,
                                path_to_excel:str
                                ) -> Tuple[List[Any], List[Tuple[Any]]]:
        #MARK: load_emails_from_excel
        wb = openpyxl.load_workbook(path_to_excel)
        ws = wb["Raw Data"]

        # Извлекаем заголовки из первой строки
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        # Извлекаем значения из остальных строк
        data = [tuple(cell.value for cell in row) for row in ws.iter_rows(min_row=2)]
        return headers, data



    def run(self):
        #MARK: run
        if self.input_excel_file:
            if self.input_excel_file.exists():
                print(f"Обнаружен входной файл Excel: {self.input_excel_file}")
                headers_excel, data_excel = self.load_emails_from_excel(path_to_excel=self.input_excel_file)
                print(f"Loaded {len(data_excel)} emails from {self.output_excel_file}")
            else:
                print(f"Обнаружен параметр входящего Excel, однако этот файл не существует: {self.input_excel_file}")
        print(f"Начинаем выгружать информацию из email. Выходной файл: {self.output_excel_file}")
        self.fetch_emails()
